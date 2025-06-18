# Written By: Brandon Johns
# Date Version Created: 2022-11-28
# Date Last Edited: 2022-11-28
# Status: functional

### PURPOSE ###
# Example for using Brandon's interface to the 'Vicon DataStream SDK'
# Output object location & orientation to excel

### NOTES ###
# Due to use of the 'multiprocessing' module,
# for some reason, any script that uses this module must be wrapped in
#   if __name__ == '__main__':
# Otherwise there will be some weird error about needing to call
#   multiprocessing.freeze_support()
# See: https://stackoverflow.com/a/24374798


import VDSInterface
import time
import numpy
import openpyxl
from datetime import datetime


# Utility for openpyxl
#   In openpyxl, the first cell is at column=1,row=1
def writeRow(sheet, row, data):
    idx = 0 # fight me
    if isinstance(data, numpy.ndarray):
        for value in data.flat:
            sheet.cell(column=idx+1, row=row, value=value)
            idx += 1
    elif isinstance(data, list) or isinstance(data, tuple):
        for value in data:
            sheet.cell(column=idx+1, row=row, value=value)
            idx += 1
    else:
        # Single value, not an array
        sheet.cell(column=idx+1, row=row, value=data)


if __name__ == "__main__":
    # Program configuration
    hostName = 'localhost:801' # IP address of the computer running Vicon Tracker
    lightWeightMode = False
    excelFilename = 'rawData_' + datetime.today().strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx'

    # Create Excel file
    # with sheets for position and orientation of each object
    workbook = openpyxl.Workbook()
    workbook.active.title = 'FrameInfo'
    workbook.create_sheet('Jac_P')
    workbook.create_sheet('Jac_R')
    workbook.create_sheet('Ped_P')
    workbook.create_sheet('Ped_R')
    # Header row
    headerFrameInfo = ['Number','Time']
    headerP = ['x', 'y', 'z']
    headerR = ['qw', 'qx', 'qy', 'qz']
    writeRow(workbook['FrameInfo'], 1, headerFrameInfo)
    writeRow(workbook['Jac_P'], 1, headerP)
    writeRow(workbook['Jac_R'], 1, headerR)
    writeRow(workbook['Ped_P'], 1, headerP)
    writeRow(workbook['Ped_R'], 1, headerR)

    # Instance the interface
    vdsi = VDSInterface.Interface()
    vdsi.Connect(hostName, lightWeightMode)

    timeStart = time.time()
    duration = 3 # [s]
    excelRow = 2 # First row has the header. Data starts at row 2
    # Collect data
    while time.time() - timeStart < duration:
        # Get next data frame
        frame = vdsi.GetFrame_GetUnread()
        Jackal = frame.GetByName('Jackal')
        Pedestrian = frame.GetByName('Pedestrian')

        # If the object is occluded, the row will be blank
        print('Frame:', frame.FrameNumber())
        writeRow(workbook['FrameInfo'], excelRow, [frame.FrameNumber(), frame.FrameTime_seconds()-timeStart])
        writeRow(workbook['Jac_P'], excelRow, Jackal.P())
        writeRow(workbook['Jac_R'], excelRow, Jackal.quat_wxyz())
        writeRow(workbook['Ped_P'], excelRow, Pedestrian.P())
        writeRow(workbook['Ped_R'], excelRow, Pedestrian.quat_wxyz())

        excelRow += 1

    vdsi.Disconnect()
    workbook.save(excelFilename)
